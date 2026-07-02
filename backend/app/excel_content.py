from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import json
import sqlite3
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.content_store import ContentNotFoundError, connect, get_admin_chapters_for_course, get_chapter


CHAPTER_COLUMNS = [
    "id",
    "course_id",
    "slug",
    "title",
    "subtitle",
    "description",
    "xp",
    "status",
    "sort_order",
]

BLOCK_COLUMNS = [
    "id",
    "chapter_id",
    "type",
    "title",
    "xp",
    "required",
    "sort_order",
    "content_json",
]

CONTENT_COLUMNS = [
    "plaats",
    "hoofdstuk_id",
    "hoofdstuk_titel",
    "blok_id",
    "blok_volgorde",
    "blok_type",
    "blok_titel",
    "veld",
    "pad",
    "content",
]


class WorkbookImportError(ValueError):
    pass


def export_course_workbook(course_id: str) -> bytes:
    chapters = [get_chapter(chapter["id"]) for chapter in get_admin_chapters_for_course(course_id)]
    if not chapters:
        raise ContentNotFoundError(f"Course has no chapters: {course_id}")

    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["Houvast content export"])
    readme.append(["Bewerk vooral de sheet Content: daar staan plaats, hoofdstuk, blok, veld en tekst leesbaar onder elkaar."])
    readme.append(["Laat kolomnamen staan. De technische sheet Blocks blijft bestaan voor complete JSON-backup."])
    readme.append(["Import leest Content terug en vervangt daarna de cursus pas na validatie."])
    readme.append(["Bij import archiveert de backend eerst de oude versie."])
    _style_sheet(readme)

    content_sheet = workbook.create_sheet("Content")
    content_sheet.append(CONTENT_COLUMNS)
    for chapter in chapters:
        for block in chapter.get("blocks", []):
            for path, value in _flatten_content(block.get("content", {})):
                content_sheet.append(
                    [
                        f"H{int(chapter.get('sort_order', 0)):02d}.B{int(block.get('sort_order', 0)):02d}",
                        chapter.get("id", ""),
                        chapter.get("title", ""),
                        block.get("id", ""),
                        int(block.get("sort_order", 1)),
                        block.get("type", ""),
                        block.get("title", ""),
                        _field_label(path),
                        path,
                        value,
                    ]
                )
    _style_sheet(content_sheet)
    content_sheet.column_dimensions["A"].width = 12
    content_sheet.column_dimensions["C"].width = 28
    content_sheet.column_dimensions["F"].width = 18
    content_sheet.column_dimensions["G"].width = 30
    content_sheet.column_dimensions["H"].width = 22
    content_sheet.column_dimensions["I"].width = 28
    content_sheet.column_dimensions["J"].width = 90

    chapter_sheet = workbook.create_sheet("Chapters")
    chapter_sheet.append(CHAPTER_COLUMNS)
    for chapter in chapters:
        chapter_sheet.append([chapter.get(column, "") for column in CHAPTER_COLUMNS])
    _style_sheet(chapter_sheet)

    block_sheet = workbook.create_sheet("Blocks")
    block_sheet.append(BLOCK_COLUMNS)
    for chapter in chapters:
        for block in chapter.get("blocks", []):
            block_sheet.append(
                [
                    block.get("id", ""),
                    chapter["id"],
                    block.get("type", ""),
                    block.get("title", ""),
                    int(block.get("xp", 0)),
                    bool(block.get("required", True)),
                    int(block.get("sort_order", 1)),
                    json.dumps(block.get("content", {}), ensure_ascii=False, indent=2),
                ]
            )
    _style_sheet(block_sheet)
    block_sheet.column_dimensions["H"].width = 90

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def import_course_workbook(course_id: str, workbook_bytes: bytes) -> dict[str, Any]:
    chapters, blocks = _read_workbook(workbook_bytes)
    _validate_import(course_id, chapters, blocks)

    now = datetime.now(timezone.utc).isoformat()
    with connect() as connection:
        current_payload = _archive_payload(connection, course_id)
        archive_cursor = connection.execute(
            """
            INSERT INTO content_archives (course_id, archived_at, reason, payload_json)
            VALUES (?, ?, ?, ?)
            """,
            (
                course_id,
                now,
                "excel_import",
                json.dumps(current_payload, ensure_ascii=False),
            ),
        )
        archive_id = archive_cursor.lastrowid
        connection.execute(
            "DELETE FROM blocks WHERE chapter_id IN (SELECT id FROM chapters WHERE course_id = ?)",
            (course_id,),
        )
        connection.execute("DELETE FROM chapters WHERE course_id = ?", (course_id,))
        for chapter in chapters:
            connection.execute(
                """
                INSERT INTO chapters (
                    id, course_id, slug, title, subtitle, description, xp, status, sort_order
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    chapter["id"],
                    chapter["course_id"],
                    chapter["slug"],
                    chapter["title"],
                    chapter["subtitle"],
                    chapter["description"],
                    chapter["xp"],
                    chapter["status"],
                    chapter["sort_order"],
                ),
            )
        for block in blocks:
            connection.execute(
                """
                INSERT INTO blocks (
                    id, chapter_id, type, title, xp, required, sort_order, content_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    block["id"],
                    block["chapter_id"],
                    block["type"],
                    block["title"],
                    block["xp"],
                    1 if block["required"] else 0,
                    block["sort_order"],
                    json.dumps(block["content"], ensure_ascii=False),
                ),
            )

    return {
        "status": "imported",
        "course_id": course_id,
        "chapters": len(chapters),
        "blocks": len(blocks),
        "archive_id": archive_id,
        "archived_at": now,
    }


def _read_workbook(workbook_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    try:
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    except Exception as exc:
        raise WorkbookImportError(f"Excel bestand kan niet gelezen worden: {exc}") from exc

    if "Chapters" not in workbook.sheetnames or "Blocks" not in workbook.sheetnames:
        raise WorkbookImportError("Excel moet sheets 'Chapters' en 'Blocks' bevatten.")

    chapters = _read_rows(workbook["Chapters"], CHAPTER_COLUMNS)
    raw_blocks = _read_rows(workbook["Blocks"], BLOCK_COLUMNS)
    blocks: list[dict[str, Any]] = []
    for index, block in enumerate(raw_blocks, start=2):
        try:
            content = json.loads(str(block.get("content_json") or "{}"))
        except json.JSONDecodeError as exc:
            raise WorkbookImportError(f"Blocks rij {index}: content_json is geen geldige JSON.") from exc
        block["content"] = content
        block.pop("content_json", None)
        blocks.append(block)
    if "Content" in workbook.sheetnames:
        _apply_content_rows(workbook["Content"], blocks)
    return chapters, blocks


def _read_rows(sheet: Any, expected_columns: list[str]) -> list[dict[str, Any]]:
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    missing = [column for column in expected_columns if column not in headers]
    if missing:
        raise WorkbookImportError(f"Sheet '{sheet.title}' mist kolommen: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        row = {column: values[headers.index(column)] for column in expected_columns}
        rows.append(_normalize_row(row))
    return rows


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    for key, value in list(row.items()):
        if isinstance(value, str):
            row[key] = value.strip()
    for key in ["xp", "sort_order"]:
        if key in row:
            row[key] = int(row.get(key) or 0)
    if "required" in row:
        value = row["required"]
        row["required"] = value is True or str(value).strip().lower() in {"true", "1", "ja", "yes"}
    return row


def _validate_import(course_id: str, chapters: list[dict[str, Any]], blocks: list[dict[str, Any]]) -> None:
    if not chapters:
        raise WorkbookImportError("Import bevat geen hoofdstukken.")

    chapter_ids: set[str] = set()
    slugs: set[str] = set()
    for index, chapter in enumerate(chapters, start=2):
        for field in ["id", "course_id", "slug", "title", "status"]:
            if not chapter.get(field):
                raise WorkbookImportError(f"Chapters rij {index}: '{field}' is verplicht.")
        if chapter["course_id"] != course_id:
            raise WorkbookImportError(f"Chapters rij {index}: course_id moet '{course_id}' zijn.")
        if chapter["id"] in chapter_ids:
            raise WorkbookImportError(f"Dubbel hoofdstuk-id: {chapter['id']}")
        if chapter["slug"] in slugs:
            raise WorkbookImportError(f"Dubbele slug: {chapter['slug']}")
        chapter_ids.add(chapter["id"])
        slugs.add(chapter["slug"])

    block_ids: set[str] = set()
    for index, block in enumerate(blocks, start=2):
        for field in ["id", "chapter_id", "type", "title"]:
            if not block.get(field):
                raise WorkbookImportError(f"Blocks rij {index}: '{field}' is verplicht.")
        if block["id"] in block_ids:
            raise WorkbookImportError(f"Dubbel block-id: {block['id']}")
        if block["chapter_id"] not in chapter_ids:
            raise WorkbookImportError(
                f"Blocks rij {index}: chapter_id '{block['chapter_id']}' bestaat niet in Chapters."
            )
        if not isinstance(block.get("content"), dict):
            raise WorkbookImportError(f"Blocks rij {index}: content_json moet een JSON-object zijn.")
        block_ids.add(block["id"])


def _flatten_content(value: Any, prefix: str = "") -> list[tuple[str, Any]]:
    if isinstance(value, dict):
        rows: list[tuple[str, Any]] = []
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten_content(child, child_prefix))
        return rows
    if isinstance(value, list):
        rows = []
        for index, child in enumerate(value):
            child_prefix = f"{prefix}[{index}]"
            rows.extend(_flatten_content(child, child_prefix))
        return rows
    return [(prefix, "" if value is None else value)]


def _field_label(path: str) -> str:
    labels = {
        "body": "Tekst",
        "summary": "Samenvatting",
        "subtitle": "Subtitel",
        "question": "Vraag",
        "prompt": "Prompt",
        "intro": "Intro",
        "case_text": "Casus",
        "callout": "Callout",
        "reference": "Bijbelgedeelte",
        "translation": "Vertaling",
        "url": "URL",
        "title": "Titel",
        "description": "Omschrijving",
        "label": "Naam",
        "default": "Standaardwaarde",
        "quote": "Quote",
        "source": "Bron",
    }
    last = path.split(".")[-1]
    if "[" in last and "]" in last:
        return last
    return labels.get(last, last.replace("_", " ").title())


def _apply_content_rows(sheet: Any, blocks: list[dict[str, Any]]) -> None:
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    required = ["blok_id", "pad", "content"]
    missing = [column for column in required if column not in headers]
    if missing:
        raise WorkbookImportError(f"Sheet 'Content' mist kolommen: {', '.join(missing)}")

    blocks_by_id = {block["id"]: block for block in blocks}
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        row = {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}
        block_id = str(row.get("blok_id") or "").strip()
        path = str(row.get("pad") or "").strip()
        if not block_id or not path:
            continue
        block = blocks_by_id.get(block_id)
        if block is None:
            raise WorkbookImportError(f"Content rij {row_number}: blok_id '{block_id}' bestaat niet in Blocks.")
        _set_content_path(block["content"], path, row.get("content"))


def _set_content_path(content: dict[str, Any], path: str, value: Any) -> None:
    parts = _parse_path(path)
    current: Any = content
    for index, part in enumerate(parts):
        last = index == len(parts) - 1
        if isinstance(part, str):
            if last:
                current[part] = "" if value is None else value
                return
            next_part = parts[index + 1]
            if part not in current or current[part] is None:
                current[part] = [] if isinstance(next_part, int) else {}
            current = current[part]
        else:
            if not isinstance(current, list):
                raise WorkbookImportError(f"Content pad '{path}' wijst naar een lijst, maar de JSON is geen lijst.")
            while len(current) <= part:
                current.append({})
            if last:
                current[part] = "" if value is None else value
                return
            current = current[part]


def _parse_path(path: str) -> list[str | int]:
    parts: list[str | int] = []
    for segment in path.split("."):
        rest = segment
        if "[" not in rest:
            parts.append(rest)
            continue
        name, remainder = rest.split("[", 1)
        if name:
            parts.append(name)
        for item in remainder.split("["):
            index_text = item.rstrip("]")
            if not index_text.isdigit():
                raise WorkbookImportError(f"Ongeldig content-pad: {path}")
            parts.append(int(index_text))
    return parts


def _archive_payload(connection: sqlite3.Connection, course_id: str) -> dict[str, Any]:
    course = connection.execute("SELECT * FROM courses WHERE id = ?", (course_id,)).fetchone()
    chapter_rows = connection.execute(
        "SELECT * FROM chapters WHERE course_id = ? ORDER BY sort_order",
        (course_id,),
    ).fetchall()
    chapters = []
    for chapter_row in chapter_rows:
        block_rows = connection.execute(
            "SELECT * FROM blocks WHERE chapter_id = ? ORDER BY sort_order",
            (chapter_row["id"],),
        ).fetchall()
        chapters.append(
            {
                **dict(chapter_row),
                "blocks": [
                    {
                        "id": block["id"],
                        "chapter_id": block["chapter_id"],
                        "type": block["type"],
                        "title": block["title"],
                        "xp": block["xp"],
                        "required": bool(block["required"]),
                        "sort_order": block["sort_order"],
                        "content": json.loads(block["content_json"]),
                    }
                    for block in block_rows
                ],
            }
        )
    return {
        "course": dict(course) if course else {"id": course_id},
        "chapters": chapters,
    }


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="20312E")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    for column in range(1, sheet.max_column + 1):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = min(max(len(str(sheet.cell(1, column).value)) + 6, 14), 48)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
